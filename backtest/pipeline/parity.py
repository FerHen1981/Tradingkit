"""Stage 1 — Pine parity engine (HARD gate).

Gate: Python vs Pine on one fixed baseline — near-equal trade count and
materially comparable win rate / PF. Parameter optimization is INVALID while
baseline parity is unresolved (pipeline v7, ground rule 1).

Two jobs live here:

1. **Properties audit** (ground rule 10). TradingView preserves previously saved
   input values when a script is replaced, so the source-code defaults are NOT
   evidence of what was tested. The exported Properties sheet is the ground
   truth, and this module compares it against the config we believe we are
   validating. A mismatch does not fail parity — it means the export tests a
   different configuration, and the comparison would be meaningless.

2. **Trade-level comparison** against the export's Trades sheet.
"""
from __future__ import annotations

from dataclasses import dataclass

# Properties label -> (config attribute, how to read the exported value).
# Keep this WIDE: a field that is not compared is a field that can silently
# differ, and stage 1 is the gate everything else rests on. The day-trail block
# is here because its absence is exactly what hid the parity gap of 24-08.
_PROP_MAP = {
    "Fixed Qty":                        ("contract_size", float),
    "Limit Order Expiry (bars)":        ("expiry_bars", int),
    "Fixed Stop (units, legacy mode)":  ("fixed_stop_ticks", float),
    "Max Stop Distance (units) — else no trade": ("max_stop_ticks", float),
    "R-Multiple (R-multiple mode)":     ("r_multiple", float),
    "Min FVG Size (units)":             ("gap_min_ticks", float),
    "Max FVG Size (units)":             ("gap_max_ticks", float),
    "Count":                            ("cvd_trend_count", int),
    "Pivot Strength (bars links/rechts)": ("pivot_k", int),
    "Stop Buffer beyond swing (units)": ("swing_buf_ticks", float),
    "Distance Unit":                    ("unit_mode", str),
    "Take-Profit Mode":                 ("tp_mode", str),
    "Drawdown Model":                   ("dd_model", str),
    "Account Phase":                    ("phase", str),
    "Day-profit exit mode":             ("day_exit_mode", str),
    "Day-trail model":                  ("day_trail_model", str),
    "Day-trail activation ($)":         ("day_trail_activation_usd", float),
    "Day-trail giveback ($)":           ("day_trail_giveback_usd", float),
    "Day-cap hard target ($)":          ("day_cap_usd", float),
    # account layer — active during the parity run, so it is compared too
    "Trailing Drawdown ($)":            ("acct_trail_dd", float),
    "PA Daily Loss Limit ($)":          ("acct_dll", float),
    "Consistency rule (%)":             ("consistency_pct", float),
    "Min. Payout ($)":                  ("min_payout", float),
    "Extra buffer above safety net ($)": ("payout_buffer", float),
}
_ONOFF_EXTRA = {
    "FVG fill check (gap invalid once mid is touched)": "use_fill_check",
    "MAE guard on (only active in Apex PA phase)": "use_mae_guard",
    "Wait-for-cap: request only at full ladder cap": "use_wait_for_cap",
}
# Exported label -> (config attribute, the string that means True)
_ENUM_AS_BOOL = {
    "Entry Mode": ("entry_limit_mode", "Limit @ 50% FVG"),
    "Stop Mode":  ("stop_swing", "Swing structure"),
}
_ONOFF = {"Use Delta Filter": "use_cvd_filter", "Use FVG Size Range Filter": "use_gap_filter",
          "Enable Break-even": "use_breakeven", "Enable Trailing": "use_trail",
          **_ONOFF_EXTRA}


@dataclass
class Export:
    path: str
    properties: dict
    trades: list      # one dict per CLOSED trade
    performance: dict

    @property
    def n_trades(self) -> int:
        return len(self.trades)

    def stats(self) -> dict:
        nets = [t["net"] for t in self.trades]
        wins = [x for x in nets if x > 0]
        losses = [x for x in nets if x <= 0]
        gl = -sum(losses)
        return {"trades": len(nets), "net": round(sum(nets), 2),
                "win_rate_pct": round(100 * len(wins) / len(nets), 1) if nets else 0.0,
                "profit_factor": round(sum(wins) / gl, 2) if gl > 0 else float("inf"),
                "longs": sum(1 for t in self.trades if t["dir"] > 0),
                "shorts": sum(1 for t in self.trades if t["dir"] < 0)}


def read_export(path: str) -> Export:
    """Parse a TradingView strategy export (.xlsx): Properties, Trades, Performance."""
    try:
        import openpyxl
    except ImportError as e:                       # a traceback here reads as a bug
        raise SystemExit(
            "trap 1 leest TradingView-exports met openpyxl, en die staat niet in "
            "deze omgeving.\n  Installeer hem in dezelfde interpreter als de lab-service:\n"
            "    /root/mex-journal/.venv-bt/bin/pip install openpyxl\n"
            "  (of: <venv>/bin/pip install -r backtest/requirements.txt)") from e
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    props = {}
    if "Properties" in wb.sheetnames:
        for row in wb["Properties"].iter_rows(values_only=True):
            if row and row[0] is not None:
                props[str(row[0]).strip()] = row[1]

    trades = []
    if "Trades" in wb.sheetnames:
        ws = wb["Trades"]
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
        ix = {name: hdr.index(name) for name in hdr}
        get = lambda r, k: (r[ix[k]] if k in ix and ix[k] < len(r) else None)
        # TradingView writes one row per LEG; pair them on trade number.
        legs: dict = {}
        for r in rows[1:]:
            num = get(r, "Trade number")
            if num is None:
                continue
            legs.setdefault(str(num), []).append(r)
        for num, rs in legs.items():
            entry = next((r for r in rs if str(get(r, "Type") or "").lower().startswith("entry")), None)
            exit_ = next((r for r in rs if str(get(r, "Type") or "").lower().startswith("exit")), None)
            if entry is None or exit_ is None:
                continue                       # still-open trade
            typ = str(get(entry, "Type") or "").lower()
            try:
                net = float(get(exit_, "Net PnL USD") or 0.0)
            except (TypeError, ValueError):
                net = 0.0
            trades.append({
                "n": int(float(num)), "dir": 1 if "long" in typ else -1,
                "entry_time": str(get(entry, "Date and time") or ""),
                "exit_time": str(get(exit_, "Date and time") or ""),
                "entry_px": _f(get(entry, "Price USD")), "exit_px": _f(get(exit_, "Price USD")),
                "qty": _f(get(entry, "Size (qty)")), "net": net,
                "signal": str(get(entry, "Signal") or ""),
                "exit_reason": str(get(exit_, "Signal") or ""),
            })
        trades.sort(key=lambda t: t["n"])

    perf = {}
    if "Performance" in wb.sheetnames:
        for row in wb["Performance"].iter_rows(values_only=True):
            if row and row[0] is not None:
                perf[str(row[0]).strip()] = row[1]
    wb.close()
    return Export(path=path, properties=props, trades=trades, performance=perf)


def _f(x, d=0.0):
    try:
        return float(x)
    except (TypeError, ValueError):
        return d


def _root(ticker) -> str:
    """CME_MINI:MES1! -> MES."""
    import re
    t = str(ticker or "").split(":")[-1].strip().upper()
    m = re.match(r"^([A-Z]+?)\d*!?$", t)
    return m.group(1) if m else t


def export_window(export: Export):
    """(start, end) datetimes of the range TradingView ACTUALLY tested.

    This is the "Backtesting range" row, not the script's own
    "Start date/time (measure from)" input — that input only suppresses entries
    and is routinely years wider than the bars TradingView loaded. Confusing the
    two is how a one-year export gets mistaken for a multi-year one."""
    from datetime import datetime
    raw = str(export.properties.get("Backtesting range") or "")
    parts = [p.strip() for p in raw.replace("\u2014", "—").split("—")]
    out = []
    for p in parts[:2]:
        for fmt in ("%b %d, %Y, %H:%M", "%b %d, %Y"):
            try:
                out.append(datetime.strptime(p, fmt))
                break
            except ValueError:
                continue
    return (out[0], out[1]) if len(out) == 2 else (None, None)


def audit_environment(export: Export, cfg) -> list[dict]:
    """Is this export even comparable to this engine on this data?

    A properties audit that only checks strategy inputs will happily report
    "0 afwijkingen" for a MES engine held against an MYM export. These are the
    checks that make the comparison meaningful at all."""
    p = export.properties
    rows = []

    def chk(label, exp, got, ok, note=""):
        rows.append({"label": label, "export": exp, "config": got,
                     "ok": bool(ok), "note": note})

    want_sym = getattr(getattr(cfg, "contract", None), "symbol", None)
    got_sym = _root(p.get("Symbol"))
    chk("Symbol", p.get("Symbol"), want_sym, got_sym == str(want_sym).upper(),
        "een export van een andere markt maakt elke vergelijking betekenisloos")

    tf = str(p.get("Timeframe") or "").strip().lower()
    chk("Timeframe", p.get("Timeframe"), "1 minute", tf in ("1 minute", "1m", "1"),
        "de vloot is op 1m gevalideerd")

    ct = getattr(cfg, "contract", None)
    if ct is not None:
        tick = _f(p.get("Tick size"), 0.0)
        chk("Tick size", tick, ct.mintick, not tick or abs(tick - ct.mintick) < 1e-9)
        pv = _f(p.get("Point value"), 0.0)
        chk("Point value", pv, ct.pointvalue, not pv or abs(pv - ct.pointvalue) < 1e-9)
        com = _f(p.get("Commission"), -1.0)
        chk("Commission", com, ct.commission_per_contract,
            com < 0 or abs(com - ct.commission_per_contract) < 1e-9,
            "kosten verschuiven PF direct — zie D-07/D-08")
        slip = _f(str(p.get("Slippage") or "").split()[0] if p.get("Slippage") else None, -1.0)
        chk("Slippage (ticks)", slip, ct.slippage_ticks,
            slip < 0 or abs(slip - ct.slippage_ticks) < 1e-9)

    prog = p.get("Firm program")
    if prog is not None:
        from . import fleet
        want = fleet.firm_program(cfg.name) if cfg.name in fleet.names() else None
        chk("Firm program", prog, want, want is None or str(prog) == want,
            "bepaalt het drawdown-model (D-20)")

    a, b = export_window(export)
    chk("Backtesting range", p.get("Backtesting range"),
        f"{a:%Y-%m-%d} -> {b:%Y-%m-%d}" if a and b else "onleesbaar", a is not None,
        "dit is het venster waarover de simulator moet draaien")
    return rows


def as_tested(cfg, audit: dict):
    """A copy of `cfg` carrying the values the export was ACTUALLY run with.

    Ground rule 10 cuts both ways. The audit's job is to say the export tests a
    different configuration than the released source — but once that is said and
    recorded, refusing to run leaves the engine itself unmeasured. This adopts
    the export's values so stage 1 can answer the question it exists for: given
    the same inputs, does the Python engine reproduce the Pine engine?

    A pass under these values is parity with THE EXPORT, not with the released
    .pine. The caller must record which it was — they are different claims, and
    the difference is exactly what Pine Dev has to decide."""
    import dataclasses
    changes = {}
    for r in audit["rows"]:
        if r["ok"]:
            continue
        attr, want = r["attr"], r["export"]
        if isinstance(want, str) and "->" in want:      # "Limit @ 50% FVG -> True"
            want = want.rsplit("->", 1)[1].strip() == "True"
        cur = getattr(cfg, attr, None)
        if isinstance(cur, bool):
            want = bool(want)
        elif isinstance(cur, int) and not isinstance(cur, bool):
            want = int(float(want))
        elif isinstance(cur, float):
            want = float(want)
        changes[attr] = (cur, want)
    if not changes:
        return cfg, {}
    return (dataclasses.replace(cfg, **{k: v[1] for k, v in changes.items()}),
            changes)


def audit_properties(export: Export, cfg) -> dict:
    """Ground rule 10 — does the export actually test the config we think it does?

    Reports coverage as well as mismatches: a Properties sheet that simply lacks
    the fields would otherwise score a clean zero, which is the most dangerous
    possible outcome for a hard gate."""
    rows, mismatches, missing = [], 0, []

    def add(label, attr, got, want, ok):
        nonlocal mismatches
        mismatches += 0 if ok else 1
        rows.append({"label": label, "attr": attr, "export": got,
                     "config": want, "ok": bool(ok)})

    for label, (attr, cast) in _PROP_MAP.items():
        if label not in export.properties:
            missing.append(label)
            continue
        raw = export.properties[label]
        have = getattr(cfg, attr, None)
        try:
            want, got = (cast(have), cast(raw)) if cast is not str else (str(have), str(raw))
        except (TypeError, ValueError):
            want, got = str(have), str(raw)
        add(label, attr, got, want, want == got)

    for label, attr in _ONOFF.items():
        if label not in export.properties:
            missing.append(label)
            continue
        got = str(export.properties[label]).strip().lower() in ("on", "true", "1")
        add(label, attr, got, bool(getattr(cfg, attr, False)),
            got == bool(getattr(cfg, attr, False)))

    for label, (attr, true_value) in _ENUM_AS_BOOL.items():
        if label not in export.properties:
            missing.append(label)
            continue
        got = str(export.properties[label]).strip() == true_value
        add(label, attr, f"{export.properties[label]} -> {got}",
            bool(getattr(cfg, attr, False)), got == bool(getattr(cfg, attr, False)))

    env = audit_environment(export, cfg)
    env_bad = [r for r in env if not r["ok"]]
    checked = len(rows)
    a, b = export_window(export)
    return {"rows": rows, "mismatches": mismatches,
            "environment": env, "environment_mismatches": len(env_bad),
            "checked": checked, "missing": missing,
            "coverage_pct": round(100 * checked / (checked + len(missing)), 1)
                            if checked + len(missing) else 0.0,
            "symbol": export.properties.get("Symbol"),
            "commission": export.properties.get("Commission"),
            "slippage": export.properties.get("Slippage"),
            "timeframe": export.properties.get("Timeframe"),
            "firm_program": export.properties.get("Firm program"),
            "window_start": a, "window_end": b,
            "start": export.properties.get("Start date/time (measure from)"),
            "end": export.properties.get("End date/time (measure through)"),
            "ok": mismatches == 0 and not env_bad and not missing}


def compare(sim_kpis: dict, export: Export,
            trade_tol: float = 0.10, pf_tol: float = 0.20, wr_tol: float = 10.0) -> dict:
    """Stage-1 verdict. 'Near-equal trade count and materially comparable WR/PF'
    is made explicit: trade count within `trade_tol`, PF within `pf_tol`
    (relative), win rate within `wr_tol` percentage points."""
    tv = export.stats()
    sim_n = int(sim_kpis.get("trades") or 0)
    tv_n = tv["trades"]
    checks = []

    def chk(name, sim, ref, ok, detail):
        checks.append({"name": name, "sim": sim, "pine": ref, "ok": bool(ok), "detail": detail})

    dn = abs(sim_n - tv_n) / tv_n if tv_n else 1.0
    chk("trade count", sim_n, tv_n, dn <= trade_tol, f"{dn*100:.1f}% apart (limit {trade_tol*100:.0f}%)")

    s_pf, t_pf = float(sim_kpis.get("profit_factor") or 0), tv["profit_factor"]
    dpf = abs(s_pf - t_pf) / t_pf if t_pf not in (0, float("inf")) else 1.0
    chk("profit factor", round(s_pf, 2), t_pf, dpf <= pf_tol, f"{dpf*100:.1f}% apart (limit {pf_tol*100:.0f}%)")

    s_wr, t_wr = float(sim_kpis.get("win_rate_pct") or 0), tv["win_rate_pct"]
    chk("win rate", s_wr, t_wr, abs(s_wr - t_wr) <= wr_tol, f"{abs(s_wr-t_wr):.1f}pp apart (limit {wr_tol:.0f}pp)")

    s_l, s_s = int(sim_kpis.get("longs") or 0), int(sim_kpis.get("shorts") or 0)
    side_ok = (tv["longs"] == 0 or abs(s_l - tv["longs"]) / max(tv["longs"], 1) <= 0.25) and \
              (tv["shorts"] == 0 or abs(s_s - tv["shorts"]) / max(tv["shorts"], 1) <= 0.25)
    chk("long/short split", f"{s_l}/{s_s}", f"{tv['longs']}/{tv['shorts']}", side_ok, "within 25% per side")

    passed = all(c["ok"] for c in checks)
    return {"pass": passed, "checks": checks, "pine": tv, "sim": {
        "trades": sim_n, "profit_factor": round(s_pf, 2), "win_rate_pct": s_wr,
        "net": sim_kpis.get("net_profit")},
        "verdict": ("parity established on this baseline" if passed else
                    "NOT at parity — fix engine semantics before any parameter search "
                    "(ground rule 1); investigate the first divergent trades, do not re-optimize")}
